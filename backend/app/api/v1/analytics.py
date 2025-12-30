# ============================================================
# 文件: backend/app/api/v1/analytics.py
# 功能: 数据分析API路由端点
# 作者: 开发团队
# 创建时间: 2025-12-20
# ============================================================

import logging
import json
import io
import csv
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from decimal import Decimal

from database.config import get_db
from app.schemas import analytics as schemas
from app.crud import pdf_file as crud
from app.services.analytics_service import AnalyticsService

logger = logging.getLogger(__name__)

router = APIRouter()


# ============================================================
# 1. 汇总统计接口
# ============================================================

@router.get(
    "/summary",
    response_model=schemas.AggregatedMetrics,
    summary="获取汇总统计数据",
    description="获取指定日期范围内的对账单汇总统计指标"
)
async def get_summary_statistics(
    start_date: date = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: date = Query(..., description="结束日期 (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """获取汇总统计数据.

    包含以下指标:
    - total_sales: 总销售额
    - total_refund: 总退款额
    - total_commission: 总佣金
    - total_wfs_fee: 总WFS费用
    - total_ads_cost: 总广告费用
    - net_revenue: 净收入
    - statement_count: 对账单数量
    - period_days: 周期天数

    示例: GET /api/v1/analytics/summary?start_date=2024-01-01&end_date=2024-12-31
    """
    logger.info(f"获取汇总统计: {start_date} 至 {end_date}")

    # 验证日期范围
    if start_date > end_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="开始日期不能晚于结束日期"
        )

    # 查询日期范围内的对账单
    statements = crud.get_statements_by_date_range(db, start_date, end_date)

    if not statements:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"在 {start_date} 至 {end_date} 期间未找到对账单数据"
        )

    # 计算汇总指标
    metrics = AnalyticsService.calculate_aggregated_metrics(statements)

    return metrics


# ============================================================
# 2. 趋势分析接口
# ============================================================

@router.get(
    "/trends",
    response_model=schemas.TrendAnalysisResponse,
    summary="获取趋势分析数据",
    description="获取指定时间粒度的对账单趋势分析数据"
)
async def get_trend_analysis(
    start_date: date = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: date = Query(..., description="结束日期 (YYYY-MM-DD)"),
    granularity: str = Query(
        "monthly",
        regex="^(monthly|weekly|statement)$",
        description="时间粒度: monthly(月), weekly(周), statement(每个对账单)"
    ),
    db: Session = Depends(get_db)
):
    """获取趋势分析数据.

    按指定时间粒度返回时间序列数据，每个时期包含汇总指标。

    Args:
        start_date: 开始日期
        end_date: 结束日期
        granularity: 时间粒度 (monthly|weekly|statement)

    示例: GET /api/v1/analytics/trends?start_date=2024-01-01&end_date=2024-12-31&granularity=monthly
    """
    logger.info(f"获取趋势分析: {start_date} 至 {end_date}, 粒度={granularity}")

    # 验证日期范围
    if start_date > end_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="开始日期不能晚于结束日期"
        )

    # 按时间粒度分组查询
    grouped_statements = crud.get_statements_grouped_by_period(
        db, start_date, end_date, granularity
    )

    if not grouped_statements:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"在 {start_date} 至 {end_date} 期间未找到对账单数据"
        )

    # 计算趋势分析
    time_series = AnalyticsService.calculate_trend_analysis(
        grouped_statements, granularity
    )

    return {
        "time_series": time_series,
        "granularity": granularity,
        "total_periods": len(time_series),
    }


# ============================================================
# 3. 对比分析接口
# ============================================================

@router.post(
    "/comparison",
    response_model=schemas.ComparisonResponse,
    summary="对比两个时间段的数据",
    description="计算两个时间段对账单数据的对比，包括绝对变化和百分比变化"
)
async def compare_periods(
    request: schemas.ComparisonRequest,
    db: Session = Depends(get_db)
):
    """对比两个时间段的对账单数据.

    比较两个时间段的财务指标，返回绝对变化量和百分比变化。

    Args:
        request: 包含两个时间段信息的请求对象

    示例:
    POST /api/v1/analytics/comparison
    {
        "period1_start": "2024-01-01",
        "period1_end": "2024-06-30",
        "period2_start": "2024-07-01",
        "period2_end": "2024-12-31",
        "granularity": "monthly"
    }
    """
    logger.info(
        f"对比分析: 第一期 {request.period1_start} 至 {request.period1_end}, "
        f"第二期 {request.period2_start} 至 {request.period2_end}"
    )

    # 验证日期范围
    if request.period1_start > request.period1_end:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="第一期: 开始日期不能晚于结束日期"
        )

    if request.period2_start > request.period2_end:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="第二期: 开始日期不能晚于结束日期"
        )

    # 查询两个时间段的对账单
    period1_statements = crud.get_statements_by_date_range(
        db, request.period1_start, request.period1_end
    )

    period2_statements = crud.get_statements_by_date_range(
        db, request.period2_start, request.period2_end
    )

    if not period1_statements and not period2_statements:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="两个时间段都未找到对账单数据"
        )

    # 计算对比
    comparison = AnalyticsService.calculate_comparison(
        period1_statements, period2_statements
    )

    return comparison


# ============================================================
# 4. 异常检测接口
# ============================================================

@router.get(
    "/anomalies",
    response_model=schemas.AnomalyDetectionResponse,
    summary="检测异常数据",
    description="检测指定日期范围内的异常对账单数据"
)
async def detect_anomalies(
    start_date: date = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: date = Query(..., description="结束日期 (YYYY-MM-DD)"),
    severity: str = Query(
        "all",
        regex="^(all|low|medium|high)$",
        description="筛选严重程度: all(全部), low(轻), medium(中), high(重)"
    ),
    db: Session = Depends(get_db)
):
    """检测异常数据.

    检测规则:
    - 高退款率: 退款/销售 > 20%
    - 负收入: 净收入 < 0
    - 高佣金率: 佣金/销售 > 20%
    - 高WFS费用: WFS费用/销售 > 10%
    - 高广告费用: 广告费用/销售 > 15%

    严重程度:
    - low: 轻微偏差
    - medium: 中等偏差
    - high: 严重偏差

    Args:
        start_date: 开始日期
        end_date: 结束日期
        severity: 筛选严重程度 (all|low|medium|high)

    示例: GET /api/v1/analytics/anomalies?start_date=2024-01-01&end_date=2024-12-31&severity=high
    """
    logger.info(f"检测异常: {start_date} 至 {end_date}, 严重程度={severity}")

    # 验证日期范围
    if start_date > end_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="开始日期不能晚于结束日期"
        )

    # 查询日期范围内的对账单
    statements = crud.get_statements_by_date_range(db, start_date, end_date)

    if not statements:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"在 {start_date} 至 {end_date} 期间未找到对账单数据"
        )

    # 检测异常
    all_anomalies = AnalyticsService.detect_anomalies(statements)

    # 按严重程度筛选
    if severity != "all":
        filtered_anomalies = [a for a in all_anomalies if a["severity"] == severity]
    else:
        filtered_anomalies = all_anomalies

    return {
        "total_statements": len(statements),
        "anomaly_count": len(filtered_anomalies),
        "anomalies": filtered_anomalies,
    }


# ============================================================
# 导出接口
# ============================================================

def _convert_decimal_to_float(obj):
    """递归转换Decimal为float，用于JSON序列化"""
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: _convert_decimal_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_convert_decimal_to_float(item) for item in obj]
    return obj


@router.get(
    "/summary/export",
    summary="导出汇总统计数据",
    description="导出指定日期范围的汇总统计数据（支持Excel/CSV/JSON格式）"
)
async def export_summary(
    start_date: date = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: date = Query(..., description="结束日期 (YYYY-MM-DD)"),
    format: str = Query("excel", regex="^(excel|csv|json)$", description="导出格式"),
    db: Session = Depends(get_db)
):
    """导出汇总统计数据"""
    logger.info(f"导出汇总统计: {start_date} 至 {end_date}, 格式={format}")

    if start_date > end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="开始日期不能晚于结束日期")

    statements = crud.get_statements_by_date_range(db, start_date, end_date)
    if not statements:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="未找到数据")

    metrics = AnalyticsService.calculate_aggregated_metrics(statements)
    metrics = _convert_decimal_to_float(metrics)

    if format == "json":
        json_data = json.dumps(metrics, ensure_ascii=False, indent=2)
        return StreamingResponse(io.BytesIO(json_data.encode()), media_type="application/json")

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(metrics.keys())
        writer.writerow(metrics.values())
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv")

    else:  # excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总统计"
        ws.append(list(metrics.keys()))
        ws.append(list(metrics.values()))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get(
    "/trends/export",
    summary="导出趋势分析数据",
    description="导出指定日期范围的趋势分析数据"
)
async def export_trends(
    start_date: date = Query(...),
    end_date: date = Query(...),
    granularity: str = Query("monthly", regex="^(monthly|weekly|statement)$"),
    format: str = Query("excel", regex="^(excel|csv|json)$"),
    db: Session = Depends(get_db)
):
    """导出趋势分析数据"""
    logger.info(f"导出趋势分析: {start_date} 至 {end_date}")

    if start_date > end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="开始日期不能晚于结束日期")

    statements = crud.get_statements_by_date_range(db, start_date, end_date)
    if not statements:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="未找到数据")

    trends = AnalyticsService.calculate_trend_analysis(statements, granularity)
    trends = _convert_decimal_to_float(trends)

    if format == "json":
        json_data = json.dumps(trends, ensure_ascii=False, indent=2, default=str)
        return StreamingResponse(io.BytesIO(json_data.encode()), media_type="application/json")

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        if trends:
            writer.writerow(trends[0].keys())
            for row in trends:
                writer.writerow(row.values())
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv")

    else:  # excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "趋势分析"
        if trends:
            ws.append(list(trends[0].keys()))
            for row in trends:
                ws.append(list(row.values()))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post(
    "/comparison/export",
    summary="导出对比分析数据",
    description="导出两个时间段的对比分析数据"
)
async def export_comparison(
    request: schemas.ComparisonRequest,
    format: str = Query("excel", regex="^(excel|csv|json)$"),
    db: Session = Depends(get_db)
):
    """导出对比分析数据"""
    logger.info(f"导出对比分析")

    period1_statements = crud.get_statements_by_date_range(db, request.period1_start, request.period1_end)
    period2_statements = crud.get_statements_by_date_range(db, request.period2_start, request.period2_end)

    if not period1_statements or not period2_statements:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="未找到数据")

    comparison = AnalyticsService.calculate_comparison(period1_statements, period2_statements)
    comparison = _convert_decimal_to_float(comparison)

    if format == "json":
        json_data = json.dumps(comparison, ensure_ascii=False, indent=2, default=str)
        return StreamingResponse(io.BytesIO(json_data.encode()), media_type="application/json")

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["指标", "期间1", "期间2", "变化量", "变化率"])
        for key in comparison.get("period1", {}).keys():
            period1_val = comparison["period1"].get(key, 0)
            period2_val = comparison["period2"].get(key, 0)
            change = comparison.get("changes", {}).get(key, {})
            writer.writerow([key, period1_val, period2_val, change.get("absolute", 0), change.get("percentage", 0)])
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv")

    else:  # excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "对比分析"
        ws.append(["指标", "期间1", "期间2", "变化量", "变化率"])
        for key in comparison.get("period1", {}).keys():
            period1_val = comparison["period1"].get(key, 0)
            period2_val = comparison["period2"].get(key, 0)
            change = comparison.get("changes", {}).get(key, {})
            ws.append([key, period1_val, period2_val, change.get("absolute", 0), change.get("percentage", 0)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get(
    "/anomalies/export",
    summary="导出异常检测数据",
    description="导出异常检测结果数据"
)
async def export_anomalies(
    start_date: date = Query(...),
    end_date: date = Query(...),
    severity: str = Query("all", regex="^(all|low|medium|high)$"),
    format: str = Query("excel", regex="^(excel|csv|json)$"),
    db: Session = Depends(get_db)
):
    """导出异常检测数据"""
    logger.info(f"导出异常检测: {start_date} 至 {end_date}")

    if start_date > end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="开始日期不能晚于结束日期")

    statements = crud.get_statements_by_date_range(db, start_date, end_date)
    if not statements:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="未找到数据")

    anomalies = AnalyticsService.detect_anomalies(statements)

    if severity != "all":
        anomalies = [a for a in anomalies if a["severity"] == severity]

    anomalies = _convert_decimal_to_float(anomalies)

    if format == "json":
        json_data = json.dumps(anomalies, ensure_ascii=False, indent=2, default=str)
        return StreamingResponse(io.BytesIO(json_data.encode()), media_type="application/json")

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        if anomalies:
            writer.writerow(anomalies[0].keys())
            for row in anomalies:
                writer.writerow(row.values())
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv")

    else:  # excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "异常检测"
        if anomalies:
            ws.append(list(anomalies[0].keys()))
            for row in anomalies:
                ws.append(list(row.values()))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
