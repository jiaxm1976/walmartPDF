# ============================================================
# 文件: backend/app/services/export_service.py
# 功能: 对账单数据导出服务（Excel、CSV、JSON格式）
# 作者: 开发团队
# 创建时间: 2025-12-20
# 说明: 支持单个导出和批量导出
# ============================================================

import json
import csv
import io
import zipfile
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Any, Optional
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config.field_aliases import get_field_alias


class ExportService:
    """对账单数据导出服务."""

    # ========== JSON导出 ==========

    @staticmethod
    def export_json(statement_data: Dict[str, Any]) -> bytes:
        """将对账单数据导出为JSON格式.

        Args:
            statement_data: 完整对账单数据（来自get_complete_statement_data）

        Returns:
            bytes: JSON格式的二进制数据
        """
        # 转换Decimal和datetime类型为JSON可序列化的类型
        def convert_for_json(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            elif isinstance(obj, (datetime, date)):
                return obj.isoformat()
            elif isinstance(obj, dict):
                return {k: convert_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert_for_json(item) for item in obj]
            return obj

        converted_data = convert_for_json(statement_data)
        json_str = json.dumps(converted_data, ensure_ascii=False, indent=2)
        return json_str.encode('utf-8')

    # ========== CSV导出 ==========

    @staticmethod
    def export_csv(statement_data: Dict[str, Any], pdf_file_info: Dict[str, Any]) -> bytes:
        """将对账单数据导出为CSV格式.

        所有数据合并为一行，列为各字段，使用中文字段别名。

        Args:
            statement_data: 完整对账单数据
            pdf_file_info: PDF文件信息

        Returns:
            bytes: CSV格式的二进制数据
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # 收集所有字段名和值
        headers = []
        values = []

        # 1. PDF文件信息字段
        for key, value in pdf_file_info.items():
            if key not in ['id', 'created_at', 'updated_at']:
                alias = get_field_alias(key, 'pdf_file')
                headers.append(alias)
                values.append(str(value))

        # 2. 各板块数据字段
        boards = [
            (statement_data['header'], 'header'),
            (statement_data['sales'], 'sales'),
            (statement_data['refund'], 'refund'),
            (statement_data['adjustment'], 'adjustment'),
            (statement_data['wfs'], 'wfs'),
            (statement_data['other_activity'], 'other_activity'),
            (statement_data['footer'], 'footer'),
            (statement_data['payment'], 'payment'),
        ]

        for board_data, section_key in boards:
            if not board_data:
                continue

            for key, value in board_data.items():
                if key not in ['id', 'pdf_file_id', 'created_at', 'updated_at']:
                    alias = get_field_alias(key, section_key)
                    headers.append(alias)
                    values.append(str(value))

        # 写入CSV：第一行是字段名，第二行是数据
        writer.writerow(headers)
        writer.writerow(values)

        result = output.getvalue()
        return result.encode('utf-8')

    # ========== Excel导出 ==========

    @staticmethod
    def export_excel(statement_data: Dict[str, Any], pdf_file_info: Dict[str, Any]) -> bytes:
        """将对账单数据导出为Excel格式.

        单个Sheet，所有数据合并为一行，列为各字段，使用中文字段别名。

        Args:
            statement_data: 完整对账单数据
            pdf_file_info: PDF文件信息

        Returns:
            bytes: Excel文件的二进制数据
        """
        wb = Workbook()
        # 删除默认的Sheet，创建新的
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        ws = wb.create_sheet('对账单数据', 0)

        # 样式定义
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 收集所有字段名和值
        headers = []
        values = []

        # 1. PDF文件信息字段
        for key, value in pdf_file_info.items():
            if key not in ['id', 'created_at', 'updated_at']:
                alias = get_field_alias(key, 'pdf_file')
                headers.append(alias)
                values.append(str(value))

        # 2. 各板块数据字段
        boards = [
            (statement_data['header'], 'header'),
            (statement_data['sales'], 'sales'),
            (statement_data['refund'], 'refund'),
            (statement_data['adjustment'], 'adjustment'),
            (statement_data['wfs'], 'wfs'),
            (statement_data['other_activity'], 'other_activity'),
            (statement_data['footer'], 'footer'),
            (statement_data['payment'], 'payment'),
        ]

        for board_data, section_key in boards:
            if not board_data:
                continue

            for key, value in board_data.items():
                if key not in ['id', 'pdf_file_id', 'created_at', 'updated_at']:
                    alias = get_field_alias(key, section_key)
                    headers.append(alias)
                    values.append(str(value))

        # 写入第一行：字段名（带样式）
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入第二行：数据值
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = value

        # 设置列宽（自动调整）
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # 冻结第一行
        ws.freeze_panes = 'A2'

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ========== 批量导出 ==========

    @staticmethod
    def export_batch_zip(
        statements_data: List[Dict[str, Any]],
        file_format: str = 'json'
    ) -> bytes:
        """将多个对账单批量导出为ZIP文件.

        Args:
            statements_data: 对账单数据列表，每项包含pdf_file_info和statement_data
            file_format: 导出格式（json/csv/excel）

        Returns:
            bytes: ZIP文件的二进制数据
        """
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for item in statements_data:
                pdf_file_info = item['pdf_file_info']
                statement_data = item['statement_data']

                filename = pdf_file_info.get('original_filename', 'statement')
                filename_base = Path(filename).stem  # 去除扩展名

                # 选择文件格式和扩展名
                if file_format == 'excel':
                    file_content = ExportService.export_excel(statement_data, pdf_file_info)
                    zip_filename = f"{filename_base}.xlsx"
                elif file_format == 'csv':
                    file_content = ExportService.export_csv(statement_data, pdf_file_info)
                    zip_filename = f"{filename_base}.csv"
                else:  # json
                    file_content = ExportService.export_json(statement_data)
                    zip_filename = f"{filename_base}.json"

                # 添加到ZIP文件
                zf.writestr(zip_filename, file_content)

        zip_buffer.seek(0)
        return zip_buffer.getvalue()

    # ========== 辅助方法 ==========

    @staticmethod
    def get_export_filename(pdf_filename: str, format: str) -> str:
        """生成导出文件名.

        Args:
            pdf_filename: 原PDF文件名
            format: 导出格式（json/csv/excel）

        Returns:
            str: 导出文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = Path(pdf_filename).stem

        if format == 'excel':
            return f"{timestamp}_{base_name}.xlsx"
        elif format == 'csv':
            return f"{timestamp}_{base_name}.csv"
        else:  # json
            return f"{timestamp}_{base_name}.json"


# ============================================================
# END OF export_service.py
# ============================================================
