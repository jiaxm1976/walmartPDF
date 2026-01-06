#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出数据到 Excel
按照输入的开始日期，导出 statements 和 section_data 数据到 Excel 文件
"""

import sqlite3
import json
import sys
import os
from openpyxl.styles import Border, Side
from datetime import datetime
from pathlib import Path
import pandas as pd
from typing import List, Dict, Any, Tuple

# 可配置：字段名映射（内部字段名 -> 显示中文）
FIELD_LABEL_MAP = {
    # 示例映射（可根据项目实际字段补充）
    '统计区间': '统计区间',
    'total_sales': '销售总额',
    'total_refunds': '退款总额',
    'net_sales': '净销售额',
}

# 默认货币符号
DEFAULT_CURRENCY = '$'

# 附件主色（近似值，可按需调整）
PRIMARY_BLUE = '2B8CE6'

# 获取项目根目录
PROJECT_ROOT = Path(__file__).resolve().parent.parent
# 默认数据库位置（优先项目 backend/data，其次项目根目录下的 data）
default_db_1 = PROJECT_ROOT / "backend" / "data" / "walmart_pdf_parser.db"
default_db_2 = PROJECT_ROOT / "data" / "walmart_pdf_parser.db"

if default_db_1.exists():
    DB_PATH = default_db_1
elif default_db_2.exists():
    DB_PATH = default_db_2
else:
    DB_PATH = default_db_1
OUTPUT_DIR = PROJECT_ROOT / "output"

# 确保可以找到数据库文件
if not DB_PATH.exists():
    print(f"错误: 数据库文件不存在: {DB_PATH}")
    sys.exit(1)


def parse_period_start_date(period_str: str) -> str:
    """
    从 statement_period 字符串中提取起始日期。
    
    Args:
        period_str: 格式为 "2025-09-06 - 2025-09-20" 的字符串
        
    Returns:
        起始日期字符串，格式为 "2025-09-06"
    """
    if not period_str or " - " not in period_str:
        return None
    start_date = period_str.split(" - ")[0].strip()
    return start_date


def parse_period_end_date(period_str: str) -> str:
    """
    从 statement_period 字符串中提取结束日期。
    Args:
        period_str: 格式为 "2025-09-06 - 2025-09-20"
    Returns:
        结束日期字符串，格式为 "2025-09-20"
    """
    if not period_str or " - " not in period_str:
        return None
    end_date = period_str.split(" - ")[1].strip()
    return end_date


def query_statements_by_date(start_date: str, end_date: str = None) -> List[Tuple[int, str, str]]:
    """
    查询开始日期 >= 指定日期的所有 statement 记录。
    
    Args:
        start_date: 格式为 "2025-09-06" 的日期字符串
        
    Returns:
        [(id, pdf_name, statement_period), ...]
    """
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, pdf_name, statement_period
        FROM statements
        ORDER BY statement_period ASC
    """)

    results = []
    for row in cursor.fetchall():
        stmt_id, pdf_name, period = row
        period_start = parse_period_start_date(period)
        period_end = parse_period_end_date(period)

        if not period_start:
            continue

        # 如果提供了 end_date，则要求与 statement 的结束日期精确匹配（保留原有严格匹配逻辑）
        if end_date is not None:
            if not period_end:
                continue
            if period_end != end_date:
                continue

        # 支持两种匹配情况：
        # 1) 目标日期落在 statement 的区间内（period_start <= start_date <= period_end）
        # 2) statement 的起始日期在目标日期之后（period_start >= start_date）
        try:
            if period_end:
                if period_start <= start_date <= period_end:
                    results.append((stmt_id, pdf_name, period))
                    continue
        except Exception:
            # 若 period_end 无法比较或格式异常，则退回到起始日期比较
            pass

        if period_start >= start_date:
            results.append((stmt_id, pdf_name, period))
    
    conn.close()
    return results


def query_section_data(stmt_ids: List[int]) -> Dict[int, List[Dict[str, Any]]]:
    """
    查询指定 statement_id 对应的所有 section_data 记录。
    
    Args:
        stmt_ids: statement ID 列表
        
    Returns:
        {stmt_id: [{'section_name': ..., 'data': {...}}, ...], ...}
    """
    if not stmt_ids:
        return {}
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 构建 SQL，使用 IN 子句
    placeholders = ",".join("?" * len(stmt_ids))
    cursor.execute(f"""
        SELECT statement_id, section_name, data
        FROM section_data
        WHERE statement_id IN ({placeholders})
        ORDER BY statement_id, ROWID
    """, stmt_ids)
    
    results = {}
    for stmt_id, section_name, data_json in cursor.fetchall():
        if stmt_id not in results:
            results[stmt_id] = []
        
        try:
            data = json.loads(data_json)
        except json.JSONDecodeError:
            data = {}
        
        results[stmt_id].append({
            'section_name': section_name,
            'data': data
        })
    
    conn.close()
    return results


def collect_all_fields(section_data_map: Dict[int, List[Dict[str, Any]]]) -> Dict[str, set]:
    """
    收集所有 section 中的所有字段名。
    
    Args:
        section_data_map: section_data 查询结果
        
    Returns:
        {'section_name': {'field1', 'field2', ...}, ...}
    """
    # 保持字段出现顺序（按记录顺序）
    fields_by_section = {}

    for stmt_id, sections in section_data_map.items():
        for section_info in sections:
            section_name = section_info['section_name']
            data = section_info['data'] or {}

            if section_name not in fields_by_section:
                fields_by_section[section_name] = []

            # 以 data 的键顺序追加字段（去重）
            for k in data.keys():
                if k not in fields_by_section[section_name]:
                    fields_by_section[section_name].append(k)

    return fields_by_section


def build_dataframe(
    statements: List[Tuple[int, str, str]],
    section_data_map: Dict[int, List[Dict[str, Any]]],
    fields_by_section: Dict[str, List[str]]
) -> pd.DataFrame:
    """
    构建 DataFrame，使用多层列索引。
    
    Args:
        statements: statement 查询结果
        section_data_map: section_data 查询结果
        fields_by_section: 各 section 的字段列表
        
    Returns:
        DataFrame with MultiIndex columns
    """
    # 构建每个 PDF (statement) 对应一行的数据结构
    rows = []

    # 动态计算导出时的板块顺序（保留发现顺序，并把特定尾部板块放到末尾）
    section_order = _compute_section_order(fields_by_section)

    # 处理每个 statement，合并其所有 section 数据到一行
    for stmt_id, pdf_name, period in statements:
        sections_for_stmt = section_data_map.get(stmt_id, [])
        # 快速映射 section_name -> data
        section_map = {s['section_name']: s['data'] for s in sections_for_stmt}

        row = {
            'PDF 名称': pdf_name,
            '对账周期': period,
        }

        # 按照 section_order 展开每个板块的字段，缺失置为 None
        for section_name in section_order:
            section_fields = fields_by_section.get(section_name, [])
            data = section_map.get(section_name, {}) or {}
            for field in section_fields:
                # 列键为 (section, field) 形式在后续构建 multiindex
                row_key = f"{section_name}_{field}"
                row[row_key] = data.get(field, None)

        # 也需要包含 fields 来自未出现在 section_order 中的其他板块
        for section_name, fields in fields_by_section.items():
            if section_name in section_order:
                continue
            data = section_map.get(section_name, {}) or {}
            for field in fields:
                row_key = f"{section_name}_{field}"
                row[row_key] = data.get(field, None)

        rows.append(row)

    # 创建 DataFrame
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # 构建列对（第一层：section 或 基本信息；第二层：字段名）
    column_pairs: List[Tuple[str, str]] = []
    # 基础列
    column_pairs.append(('基本信息', 'PDF 名称'))
    column_pairs.append(('基本信息', '对账周期'))

    # 先按指定顺序追加列
    for section_name in section_order:
        section_fields = fields_by_section.get(section_name, [])
        for field in section_fields:
            column_pairs.append((section_name, field))

    # 再追加剩余板块
    for section_name, section_fields in fields_by_section.items():
        if section_name in section_order:
            continue
        for field in section_fields:
            column_pairs.append((section_name, field))

    # 使用 column_pairs 从 df 构造最终的字典数据（确保列顺序与 multi-header 一致）
    new_data = {}
    for section, field in column_pairs:
        if section == '基本信息':
            col_name = field
            new_data[(section, field)] = df.get(col_name, pd.Series([None] * len(df)))
        else:
            col_name = f"{section}_{field}"
            new_data[(section, field)] = df.get(col_name, pd.Series([None] * len(df)))

    # 创建最终 DataFrame（MultiIndex 列）
    df_result = pd.DataFrame(new_data)

    return df_result


def _translate_field(field: str) -> str:
    """把内部字段名翻译为中文展示名，若未提供映射则做回退格式化。"""
    if not field:
        return field
    if field in FIELD_LABEL_MAP:
        return FIELD_LABEL_MAP[field]
    # 如果字段已经包含中文则直接返回
    try:
        if any('\u4e00' <= ch <= '\u9fff' for ch in field):
            return field
    except Exception:
        pass
    # 回退：snake_case -> Title Case（空格分隔）
    return field.replace('_', ' ').strip().title()


def _autosize_columns_and_rows(ws, min_col=1, max_col=None):
    """为给定 sheet 自动调整列宽并根据内容估算行高。
    - 计算每列最大字符长度并设置列宽（限制最小/最大宽度）
    - 对每行估算行高：根据行内最长文本与该列宽估算换行数，设置合适的高度
    """
    from openpyxl.utils import get_column_letter

    if max_col is None:
        max_col = ws.max_column

    # 计算每列的最大文本长度
    col_max_len = {}
    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            try:
                l = len(str(cell.value))
            except Exception:
                l = 0
            if l > max_len:
                max_len = l
        col_max_len[col_idx] = max_len

    # 设置列宽：根据最大字符数，限制最小为8，最大为60
    for col_idx, max_len in col_max_len.items():
        col_letter = get_column_letter(col_idx)
        width = max(8, min(int(max_len * 1.1) + 2, 60))
        ws.column_dimensions[col_letter].width = width

    # 估算每行高度：找到该行中按列宽换行后的最大行数
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            text = str(cell.value)
            # 近似每行字符数 = 列宽 * 1 (字符近似)
            col_letter = get_column_letter(col_idx)
            col_width = ws.column_dimensions[col_letter].width or 10
            if col_width <= 0:
                col_width = 10
            # 估算需要的行数
            lines = max(1, int(len(text) / (col_width * 1.0)) + 1)
            if lines > max_lines:
                max_lines = lines
        # 设置行高，基线18，按行数增加，每行+14，高度限制为 120
        height = min(120, max(18, 18 + (max_lines - 1) * 14))
        ws.row_dimensions[row_idx].height = height


def _invert_hex_color(hex_color: str) -> str:
    """返回给定 6 位 hex 颜色的反色（不带 #）。
    接受格式 'RRGGBB' 或 'AARRGGBB'，返回 'RRGGBB'。
    """
    if not hex_color:
        return 'FFFFFF'
    s = hex_color.strip().lstrip('#')
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return 'FFFFFF'
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
    except Exception:
        return 'FFFFFF'
    ir = 255 - r
    ig = 255 - g
    ib = 255 - b
    return f"{ir:02X}{ig:02X}{ib:02X}"



def _compute_section_order(fields_by_section: Dict[str, List[str]]) -> List[str]:
    """根据实际字段字典计算导出时的板块顺序：
    - 保持 `fields_by_section` 的插入/发现顺序（即记录顺序）
    - 如果存在，保证尾部顺序为 ['其他活动','footer','right_section']
    """
    tail = ['其他活动', 'footer', 'right_section']
    sections = list(fields_by_section.keys())

    order = []
    # 如果有 header，优先放在最前面
    if 'header' in sections:
        order.append('header')

    for s in sections:
        if s == 'header' or s in tail:
            continue
        order.append(s)

    # 最后按指定尾部顺序追加（仅当存在时）
    for t in tail:
        if t in sections:
            order.append(t)

    return order


def export_to_excel(
    start_date: str,
    end_date: str = None,
    output_file: str = None
) -> str:
    """
    导出数据到 Excel 文件。
    
    Args:
        start_date: 开始日期，格式为 "2025-09-06"
        output_file: 输出文件路径，如果不指定则使用默认名称
        
    Returns:
        输出文件路径
    """
    print(f"开始导出数据，起始日期: {start_date}")
    print()
    
    # 验证日期格式
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
    except ValueError:
        print(f"错误: 日期格式不正确，应为 'YYYY-MM-DD'，您输入的是 '{start_date}'")
        sys.exit(1)
    
    # 1. 查询 statements
    print("步骤 1/4: 查询 statements 表...")
    statements = query_statements_by_date(start_date, end_date)
    print(f"  ✓ 找到 {len(statements)} 条记录")
    
    if not statements:
        print("警告: 没有找到符合条件的数据")
        return None
    
    stmt_ids = [row[0] for row in statements]
    
    # 2. 查询 section_data
    print("步骤 2/4: 查询 section_data 表...")
    section_data_map = query_section_data(stmt_ids)
    total_sections = sum(len(sections) for sections in section_data_map.values())
    print(f"  ✓ 找到 {total_sections} 条 section 记录")
    
    # 3. 收集字段
    print("步骤 3/4: 收集所有字段...")
    fields_by_section = collect_all_fields(section_data_map)
    total_fields = sum(len(fields) for fields in fields_by_section.values())
    print(f"  ✓ 找到 {len(fields_by_section)} 个板块，共 {total_fields} 个字段")
    for section, fields in sorted(fields_by_section.items()):
        print(f"    - {section}: {len(fields)} 个字段")
    
    # 4. 构建 DataFrame
    print("步骤 4/4: 构建 Excel 数据...")
    df = build_dataframe(statements, section_data_map, fields_by_section)
    print(f"  ✓ DataFrame 尺寸: {df.shape[0]} 行 × {df.shape[1]} 列")
    
    # 5. 保存到 Excel
    if output_file is None:
        # 使用日期作为文件名
        date_str = start_date.replace("-", "")
        output_file = OUTPUT_DIR / f"数据导出_{date_str}.xlsx"
    else:
        output_file = Path(output_file)
    
    # 确保输出目录存在
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    print()
    print(f"保存到 Excel 文件: {output_file}")
    
    # 手动创建 Excel 文件，支持两层表头
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = '数据导出'
    
    # 获取列结构
    if isinstance(df.columns, pd.MultiIndex):
        # 多层列索引
        col_pairs = list(df.columns)
    else:
        # 单层列索引，转换为多层
        col_pairs = [(section, col) for section, col in [('基本信息', c) if '基本信息_' not in str(c) else (c.split('_')[0], c.split('_', 1)[1]) for c in df.columns]]
    
    # 写入表头（第一行：section_name；第二行：field） - 使用简洁、清晰的浅色样式与细边框
    row = 1
    col = 1
    first_row = []
    second_row = []

    for section, field in col_pairs:
        first_row.append(section)
        second_row.append(field)

    # 样式定义（简洁商务风格）
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 一级表头：使用附件主色（浅底）+ 深蓝字
    header_fill_1 = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type='solid')
    header_font_1 = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    # 二级表头：白底 + 深蓝字（细边框，层级分明）
    header_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_font_2 = Font(bold=True, color='0F4C81', size=11, name='Calibri')

    # 第一行（先写入值，后合并相邻相同内容的单元格并统一样式）
    for col_idx, section_name in enumerate(first_row, 1):
        c = ws.cell(row=1, column=col_idx, value=section_name)
        c.font = header_font_1
        c.fill = header_fill_1
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    # 合并相邻且内容相同的第一行单元格
    col_count = len(first_row)
    i = 0
    while i < col_count:
        j = i
        while j + 1 < col_count and first_row[j + 1] == first_row[i]:
            j += 1

        start_col = i + 1
        end_col = j + 1
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            # 清除合并后除首格之外的单元格值（保持整洁）
            for clear_col in range(start_col + 1, end_col + 1):
                ws.cell(row=1, column=clear_col, value=None)

        i = j + 1

    # 第二行（字段名 -> 翻译后的中文或回退格式）
    for col_idx, field_name in enumerate(second_row, 1):
        display_name = _translate_field(field_name)
        c = ws.cell(row=2, column=col_idx, value=display_name)
        c.font = header_font_2
        c.fill = header_fill_2
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    
    # 写入数据行（并应用交替行色与细边框）
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alt_fill2 = PatternFill(start_color='FBFBFB', end_color='FBFBFB', fill_type='solid')

    for data_row_idx, (_, row_data) in enumerate(df.iterrows(), 3):
        row_fill = alt_fill if (data_row_idx % 2 == 1) else alt_fill2
        for col_idx, (section, field) in enumerate(col_pairs, 1):
            # 从 DataFrame 获取值
            if isinstance(df.columns, pd.MultiIndex):
                value = row_data[(section, field)]
            else:
                # 单层列，需要手动查找
                if section == '基本信息':
                    value = row_data[field]
                else:
                    col_name = f"{section}_{field}"
                    value = row_data[col_name] if col_name in row_data.index else None

            cell = ws.cell(row=data_row_idx, column=col_idx, value=value)

            # 数值（美元）处理：统一为两位小数并显示货币符号，且右对齐
            is_number = False
            if isinstance(value, (int, float)):
                is_number = True
            else:
                # 尝试从字符串中解析数字（去掉逗号与货币符号）
                if isinstance(value, str):
                    sval = value.strip().replace(',', '').replace('$', '').replace('¥', '')
                    try:
                        fval = float(sval)
                        cell.value = fval
                        is_number = True
                    except Exception:
                        is_number = False

            if is_number:
                try:
                    cell.number_format = f'"{DEFAULT_CURRENCY}"#,##0.00'
                except Exception:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            cell.border = border
            cell.fill = row_fill
    
    # 设置列宽
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(col_pairs) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # 计算最大宽度
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结前两行（表头）
    ws.freeze_panes = "A3"

    # 应用主表自动列宽与行高调整
    _autosize_columns_and_rows(ws, min_col=1, max_col=len(col_pairs))

    # 设置表头行高以适配阅读
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    # 为数据行设置最小行高（提高可读性）
    for r in range(3, ws.max_row + 1):
        if not ws.row_dimensions[r].height:
            ws.row_dimensions[r].height = 18

    # 外框加粗（表头与数据区域外围使用中等粗线），内部使用细线
    max_row = ws.max_row
    max_col = len(col_pairs)
    thin_side = Side(border_style='thin', color='DDDDDD')
    outer_side = Side(border_style='medium', color='0F4C81')

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            top = thin_side
            bottom = thin_side
            left = thin_side
            right = thin_side

            if row_idx == 1:
                top = outer_side
            if row_idx == max_row:
                bottom = outer_side
            if col_idx == 1:
                left = outer_side
            if col_idx == max_col:
                right = outer_side

            ws.cell(row=row_idx, column=col_idx).border = Border(left=left, right=right, top=top, bottom=bottom)

    # 为每个 PDF 创建单独的 sheet（从第二个 sheet 开始）
    # 每个 sheet 只包含两列：第一列为板块（相同板块的相邻行将合并），第二列为明细（字段名: 值）
    from openpyxl.utils import get_column_letter

    # 生成板块顺序（与主表一致，基于 fields_by_section 的发现顺序）
    section_order = _compute_section_order(fields_by_section)

    # 帮助函数：清理 sheet 名称并保证唯一性
    def _safe_sheet_name(name: str, existing: set) -> str:
        # 移除或替换 Excel 不允许的字符，限制长度 31
        safe = str(name).strip()
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            safe = safe.replace(ch, '_')
        if len(safe) > 28:
            safe = safe[:28]
        base = safe
        i = 1
        while safe in existing:
            safe = f"{base}_{i}"
            i += 1
        existing.add(safe)
        return safe

    existing_names = {wb.sheetnames[0]}

    # 使用原始 statements 和 section_data_map 来构建每个 sheet 的行
    # 我们希望第一个 PDF sheet (workbook sheet index 1) 的文件名显示为与背景相反的颜色
    for idx, (stmt_id, pdf_name, period) in enumerate(statements, start=1):
        sheet_title = _safe_sheet_name(pdf_name, existing_names)
        ws2 = wb.create_sheet(title=sheet_title)

        # 第一行标题（PDF 名称和周期）
        ws2.cell(row=1, column=1, value="文件名")
        ws2.cell(row=1, column=2, value=pdf_name)
        ws2.cell(row=2, column=1, value="对账周期")
        ws2.cell(row=2, column=2, value=period)

        # 应用主表相同的双层表头配色与字体样式（保持一致性）
        # 一级表头使用 header_fill_1/header_font_1，二级使用 header_fill_2/header_font_2
        for col_idx in (1, 2):
            c1 = ws2.cell(row=1, column=col_idx)
            c1.fill = header_fill_1
            c1.font = header_font_1
            c1.alignment = Alignment(horizontal='center', vertical='center')
            c1.border = border

            c2 = ws2.cell(row=2, column=col_idx)
            c2.fill = header_fill_2
            c2.font = header_font_2
            c2.alignment = Alignment(horizontal='center', vertical='center')
            c2.border = border

        # 对每个 sheet，将文件名单元格的字体色设置为背景反色（避免与背景同色）
        fn_cell = ws2.cell(row=1, column=2)
        # 尝试读取单元格填充的实际颜色并计算反色
        try:
            fill_color = getattr(fn_cell.fill.start_color, 'rgb', None) or getattr(fn_cell.fill.start_color, 'index', None) or PRIMARY_BLUE
            # 有些颜色可能带 ARGB 前缀，去掉 FF 前缀
            if isinstance(fill_color, str) and fill_color.startswith('FF') and len(fill_color) == 8:
                hex_in = fill_color[2:]
            else:
                hex_in = str(fill_color)
            inv = _invert_hex_color(hex_in)
            inv_argb = f"FF{inv}"
            fn_cell.font = Font(bold=True, color=inv_argb, size=11, name='Calibri')
        except Exception:
            fn_cell.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')

        # 从第三行开始写入具体明细
        write_row = 3

        sections_for_stmt = section_data_map.get(stmt_id, [])

        # 按记录顺序构建 rows_for_sheet：使用 sections_for_stmt 的顺序，并按 data 的键顺序写字段
        rows_for_sheet = []
        for s in sections_for_stmt:
            sname = s['section_name']
            data = s.get('data') or {}
            for field, val in data.items():
                rows_for_sheet.append((sname, field, val))

        # 确保“其他活动”在“footer”之前（若两者同时存在）
        section_names = [r[0] for r in rows_for_sheet]
        if '其他活动' in section_names and 'footer' in section_names:
            other_rows = [r for r in rows_for_sheet if r[0] == '其他活动']
            rows_for_sheet = [r for r in rows_for_sheet if r[0] != '其他活动']
            insert_at = next((i for i, r in enumerate(rows_for_sheet) if r[0] == 'footer'), len(rows_for_sheet))
            rows_for_sheet[insert_at:insert_at] = other_rows

        # 写入字段和值列（第2列：字段名；第3列：值）
        if rows_for_sheet:
            for idx, (section_name, field, value) in enumerate(rows_for_sheet, start=write_row):
                display_field = _translate_field(field)
                cell_field = ws2.cell(row=idx, column=2, value=display_field)
                cell_field.font = Font(bold=False, color='1F3B4D', size=11, name='Calibri')
                cell_field.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                cell_value = ws2.cell(row=idx, column=3, value=None)
                if value is None:
                    cell_value.value = None
                else:
                    # 解析数值并格式化为带货币符号的两位小数
                    if isinstance(value, (int, float)):
                        cell_value.value = value
                        try:
                            cell_value.number_format = f'"{DEFAULT_CURRENCY}"#,##0.00'
                        except Exception:
                            cell_value.number_format = '#,##0.00'
                    else:
                        sval = str(value).strip()
                        sval2 = sval.replace(',', '').replace('$', '').replace('¥', '')
                        try:
                            f = float(sval2)
                            cell_value.value = f
                            try:
                                cell_value.number_format = f'"{DEFAULT_CURRENCY}"#,##0.00'
                            except Exception:
                                cell_value.number_format = '#,##0.00'
                        except Exception:
                            cell_value.value = sval

                cell_value.alignment = Alignment(horizontal='right', vertical='center')

            # 合并第一列中连续相同的 section
            cur_section = rows_for_sheet[0][0]
            span_start = write_row
            for offset, (section_name, _, _) in enumerate(rows_for_sheet, start=write_row):
                if section_name != cur_section:
                    span_end = offset - 1
                    if span_end >= span_start:
                        if span_end > span_start:
                            ws2.merge_cells(start_row=span_start, start_column=1, end_row=span_end, end_column=1)
                        cell = ws2.cell(row=span_start, column=1)
                        cell.value = cur_section
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    cur_section = section_name
                    span_start = offset

            # 处理最后一组
            final_end = write_row + len(rows_for_sheet) - 1
            if final_end >= span_start:
                if final_end > span_start:
                    ws2.merge_cells(start_row=span_start, start_column=1, end_row=final_end, end_column=1)
                cell = ws2.cell(row=span_start, column=1)
                cell.value = cur_section
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动列宽（三列）
        # 使用自适应列宽和行高函数
        _autosize_columns_and_rows(ws2, min_col=1, max_col=3)

        # 样式微调：简洁清晰的大气风格
        header_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
        ws2.cell(row=1, column=1).font = Font(bold=True)
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=2, column=1).fill = header_fill
        ws2.cell(row=2, column=2).fill = header_fill

        # 给字段/值列加细边框与交替行背景，增强可读性
        thin = Side(border_style='thin', color='E5E7EB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_a = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        fill_b = PatternFill(start_color='FCFCFD', end_color='FCFCFD', fill_type='solid')
        for r in range(3, ws2.max_row + 1):
            f = fill_a if (r % 2 == 1) else fill_b
            for c in (1, 2, 3):
                cell = ws2.cell(row=r, column=c)
                cell.border = border
                cell.fill = f

        ws2.freeze_panes = "A3"

    # 保存工作簿到文件
    try:
        wb.save(str(output_file))
        print(f"  ✓ 已写入文件: {output_file}")
    except Exception as e:
        print(f"错误: 保存 Excel 文件失败: {e}")
        return None

    return str(output_file)


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法: python export_data_to_excel.py <start_date> [end_date] [output_file]")
        print()
        print("参数说明:")
        print("  start_date: 开始日期，格式为 YYYY-MM-DD (例如: 2025-09-06)")
        print("  end_date: (可选) 结束日期，若提供则必须与 PDF 的统计结束日期严格匹配")
        print("  output_file: (可选) 输出 Excel 文件路径")
        print()
        print("示例:")
        print("  python export_data_to_excel.py 2025-09-06")
        print("  python export_data_to_excel.py 2025-09-06 2025-09-20")
        print("  python export_data_to_excel.py 2025-09-06 2025-09-20 output.xlsx")
        sys.exit(1)

    start_date = sys.argv[1]
    end_date = None
    output_file = None
    if len(sys.argv) == 3:
        # 可能是 end_date 或 output_file（通过扩展名推断）
        if sys.argv[2].endswith('.xlsx') or sys.argv[2].endswith('.xls'):
            output_file = sys.argv[2]
        else:
            end_date = sys.argv[2]
    if len(sys.argv) >= 4:
        end_date = sys.argv[2]
        output_file = sys.argv[3]

    export_to_excel(start_date, end_date, output_file)


if __name__ == "__main__":
    main()
